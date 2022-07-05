#! /usr/bin/env python
# -*- coding: utf-8 -*-

import csv
from metrica import metrica
import gspread
from oauth2client.service_account import ServiceAccountCredentials

from openpyxl.styles import PatternFill

CREDENTIALS_FILE = 'phonebot-114ebe93cc30.json'
AD_COAST = "ad_coast"
BOUNCES = "bounces"
GOALS = "goals"
EFFICIENCY = "efficiency"
CR = "CR"
CPA = "CPA"

opt_symbol = "#"
ef_BOUNCES = opt_symbol + "Отказы"
ef_CR = opt_symbol + "CR"
ef_CPA = opt_symbol + "CPA"


class metmain(object):
    def __init__(self, met_id, token, targets=[], dir_client=None):
        self.met = metrica(met_id, token, targets, dir_client)
        credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE,
                                                                       ['https://www.googleapis.com/auth/drive'])
        self.client = gspread.authorize(credentials)
        self.targets = targets
        self.errors = []

    def make_data(self, settings, dates, bytime=False):
        try:
            data = self.met.get(settings, dates)
            return data
        except:
            self.errors.extend(self.met.errors)
            return []

    def make_data_bytime(self, settings, dates):
        try:
            data = self.met.get_bytime(settings, dates)
            return data
        except:
            self.errors.extend(self.met.errors)
            return []

    def make_file(self, list, file='documents/out.csv'):
        with open(file, 'w') as myfile:
            wr = csv.writer(myfile, quoting=csv.QUOTE_ALL)
            for row in list:
                wr.writerow(row)

    def make_list(self, data):
        list = []

        new_list = []
        for key in data[0]['dimensions']:
            new_list.append(key)

        for key in data[0]['metrics']:
            new_list.append(key)

        if EFFICIENCY in data[0]:
            for key in data[0][EFFICIENCY]:
                new_list.append(key)

        list.append(new_list)

        for row in data:
            new_list = []
            for key in row['dimensions']:
                new_list.append(row['dimensions'][key])

            for key in row['metrics']:
                new_list.append(row['metrics'][key])

            if EFFICIENCY in row:
                for key in row[EFFICIENCY]:
                    new_list.append(row[EFFICIENCY][key])

            list.append(new_list)

        return list

    def calc_efficiency(self, data):

        new_data = data
        have_goals = False
        have_ad_coast = False

        sum_goals = 0
        sum_bounces = 0
        sum_coast = 0
        sum_visits = 0

        counter = 0
        for row in data:
            efficiency = {}
            visits = 0
            goals = 0
            bounces_rate = 0
            ad_cost = 0

            for key in row['metrics']:
                if 'visits' in key:
                    visits = row['metrics'][key]
                    sum_visits += row['metrics'][key]

                if 'bounceRate' in key:
                    bounces_rate = row['metrics'][key]

                if 'goal' in key:
                    have_goals = True
                    goals += row['metrics'][key]
                    sum_goals += row['metrics'][key]

                if 'AdCost' in key:
                    have_ad_coast = True
                    ad_cost = row['metrics'][key]
                    sum_coast += row['metrics'][key]

            efficiency[BOUNCES] = int(visits * bounces_rate)
            sum_bounces += efficiency[BOUNCES]
            if have_goals:
                efficiency[GOALS] = int(goals)
                efficiency[CR] = goals / visits
            if have_ad_coast:
                efficiency[ad_cost] = ad_cost
            if have_goals and have_ad_coast:
                efficiency[CPA] = ad_cost / goals

            new_data[counter][EFFICIENCY] = efficiency
            counter += 1

        avr_CPA = 0
        avr_CR = 0

        if have_goals:
            avr_CR = sum_goals / sum_visits
            if have_ad_coast:
                avr_CPA = sum_coast / sum_goals

        for i in new_data:
            br = 0
            for key in i['metrics']:
                if 'bounceRate' in key:
                    br = i['metrics'][key]

            i[EFFICIENCY][ef_BOUNCES] = self.calc_coefficient(sum_bounces / sum_visits, br)
            if have_goals:
                i[EFFICIENCY][ef_CR] = self.calc_coefficient(i[EFFICIENCY][CR], avr_CR)
                if have_ad_coast:
                    i[EFFICIENCY][ef_CPA] = self.calc_coefficient(i[EFFICIENCY][CPA], avr_CPA)

        out_data = new_data
        return out_data

    def calc_efficiency_by_group(self, data, dems, name):
        new_data = data
        have_goals = False
        have_ad_coast = False
        dimensions_row = list(data[0]['dimensions'])

        sum_visits = {}
        sum_bounces = {}
        sum_goals = {}
        sum_coast = {}

        for row in data:

            avr_dict = {}
            dk = self.dem_key(row, dems)
            vis = 0
            bounces_rate = 0
            for key in row['metrics']:
                if 'visits' in key:
                    if dk in sum_visits:
                        sum_visits[dk] += row['metrics'][key]
                    else:
                        sum_visits[dk] = row['metrics'][key]
                    vis = row['metrics'][key]

                if 'bounceRate' in key:
                    bounces_rate = row['metrics'][key]

                if 'goal' in key:
                    have_goals = True
                    if dk in sum_goals:
                        sum_goals[dk] += row['metrics'][key]
                    else:
                        sum_goals[dk] = row['metrics'][key]

                if 'AdCost' in key:
                    have_ad_coast = True
                    if dk in sum_coast:
                        sum_coast[dk] += row['metrics'][key]
                    else:
                        sum_coast[dk] = row['metrics'][key]

                if dk in sum_bounces:
                    sum_bounces[dk] += vis * bounces_rate
                else:
                    sum_bounces[dk] = vis * bounces_rate

        avr_CPA = {}
        avr_CR = {}

        if have_goals:
            for key in sum_goals:
                if sum_visits[key] != 0:
                    avr_CR[key] = sum_goals[key] / sum_visits[key]
                else:
                    avr_CR[key] = 0
                avr_dict[GOALS] = sum_goals[key]

            if have_ad_coast:
                for key in sum_goals:
                    if sum_goals[key] != 0:
                        avr_CPA[key] = sum_coast[key] / sum_goals[key]
                    else:
                        avr_CPA[key] = 0

        for i in new_data:
            dk = self.dem_key(i, dems)
            br = 0
            for key in i['metrics']:
                if 'bounceRate' in key:
                    br = i['metrics'][key]
            if br != 0:
                i[EFFICIENCY][name + ef_BOUNCES] = self.calc_coefficient(sum_bounces[dk] / sum_visits[dk], br)
            else:
                i[EFFICIENCY][name + ef_BOUNCES] = 0
            if have_goals:
                i[EFFICIENCY][name + ef_CR] = self.calc_coefficient(i[EFFICIENCY][CR], avr_CR[dk])
                if have_ad_coast:
                    i[EFFICIENCY][name + ef_CPA] = self.calc_coefficient(avr_CPA[dk], i[EFFICIENCY][CPA])

        out_data = new_data

        return out_data

    def cals_e_baisic(self, data):
        new_data = data
        have_goals = False
        have_ad_coast = False

        counter = 0
        for row in data:
            efficiency = {}
            visits = 0
            goals = 0
            bounces_rate = 0
            ad_cost = 0

            for key in row['metrics']:
                if 'visits' in key:
                    visits = row['metrics'][key]

                if 'bounceRate' in key:
                    bounces_rate = row['metrics'][key] / 100

                if 'goal' in key:
                    have_goals = True
                    goals += row['metrics'][key]

                if 'AdCost' in key:
                    have_ad_coast = True
                    ad_cost = row['metrics'][key]

            efficiency[BOUNCES] = int(visits * bounces_rate)
            if have_goals:
                efficiency[GOALS] = int(goals)
                if visits != 0:
                    efficiency[CR] = goals / visits
                else:
                    efficiency[CR] = 0
            if have_ad_coast:
                efficiency['AdCost'] = ad_cost
                if visits != 0:
                    efficiency['CPC'] = ad_cost/visits
                else:
                    efficiency['CPC'] = 0
            if have_goals and have_ad_coast:
                if goals != 0:
                    efficiency[CPA] = ad_cost / goals
                else:
                    efficiency[CPA] = 0

            new_data[counter][EFFICIENCY] = efficiency
            counter += 1

        return new_data

    def calc_coefficient(self, val1, val2):
        val = 0
        if val2 != 0:
            if val1 > val2:
                val = "+{:0.2f}%".format((val1 / val2 - 1) * 100)
            elif val1 < val2:
                val = "-{:0.2f}%".format((1 - val1 / val2) * 100)
        return val

    def assing_dems(self, dems):
        str_out = "@"
        for i in dems:
            if i != None:
                str_out += i + "###"
            else:
                str_out += "None" + "###"
        return str_out[:-3]

    def divide_dems(self, str_in):
        return str_in.replace("@", "").split("###")

    def get_dems(self, row, dems):
        demetions = list(row['dimensions'].keys())
        lst = []
        for i in dems:
            lst.append(row['dimensions'][demetions[i]])
        return lst

    def dem_key(self, row, dems):
        return self.assing_dems(self.get_dems(row, dems))

    def sum_days(self, data):
        new_data = []
        for row in data:
            was_break = False
            if len(new_data) == 0:
                new_data.append(row)

            for new_dict in new_data:
                if row['dimensions'] == new_dict['dimensions']:
                    for key in new_dict['metrics']:
                        new_dict['metrics'][key] += row['metrics'][key]
                    for key in new_dict['efficiency']:
                        new_dict['efficiency'][key] += row['efficiency'][key]
                    was_break = True
                    break

            if not was_break:
                new_data.append(row)

        visits_key = ''
        bounce_key = ''
        adcoast_key = ''

        for key in new_data[0]['metrics']:
            if 'bounceRate' in key:
                bounce_key = key
            if 'visits' in key:
                visits_key = key
            if 'AdCost' in key:
                adcoast_key = key

        for row in new_data:
            row['dimensions']['day'] = self.get_day(row['dimensions']['day'])
            if row['metrics'][visits_key] != 0:
                row['metrics'][bounce_key] = row['efficiency']['bounces'] / row['metrics'][visits_key]
            else:
                row['metrics'][bounce_key] = 0
            if 'CR' in row['efficiency']:
                if row['metrics'][visits_key] != 0:
                    row['efficiency']['CR'] = row['efficiency']['goals'] / row['metrics'][visits_key]
                else:
                    row['efficiency']['CR'] = 0
            if 'CPA' in row['efficiency']:
                if row['efficiency']['goals'] != 0:
                    row['efficiency']['CPA'] = row['metrics'][adcoast_key] / row['efficiency']['goals']
                else:
                    row['efficiency']['CPA'] = 0

        return new_data

    def get_day(self, num):
        day = ''
        if num == '0':
            day = "7. Воскресенье"
        if num == '1':
            day = "1. Понедельник"
        if num == '2':
            day = "2. Вторник"
        if num == '3':
            day = "3. Среда"
        if num == '4':
            day = "4. Четверг"
        if num == '5':
            day = "5. Пятница"
        if num == '6':
            day = "6. Суббота"
        return day

    def make_page(self, data, name, _wb, sheet=None):
        if sheet is None:
            wb = _wb.create_sheet()
        else:
            wb = sheet

        wb.title = name

        del_list = ['ym:s:', 'ym:ad:', '<attribution>', '<currency>']
        rep_dict = {'DirectClickOrder': 'РК Директа', 'DirectBannerGroup': 'Группа объявлений',
                    'DirectPhraseOrCond': 'Ключевое слово', 'visits': 'Визиты', 'day': 'День недели',
                    'pageviews': 'Просмотры страниц', 'users': 'Пользователи',
                    'bounceRate': 'Процент оказов', 'goals': 'Заявки', 'bounces': 'Отказы', 'AdCost': 'Расход',
                    'DirectClickBanner': 'Объявление', 'DirectOrder': 'РК Директа', 'ageInterval': 'Возраст',
                    'gender': 'Пол', 'SourceEngine': 'Источник трафика', 'DirectBanner': 'Объявление',
                    'lastSignDirectClickOrder': 'РК Директа', 'hour': 'Час',
                    'DirectPlatformType': 'Тип площадки', 'DirectPlatform': 'Площадка', "lastSignDirectOrder":'РК Директа'}

        for i in range(len(data[0])):
            for j in del_list:
                data[0][i] = str(data[0][i]).replace(j, "")

        for i in range(len(data[0])):
            for key in rep_dict:
                if key == data[0][i]:
                    data[0][i] = rep_dict[key]
            if 'AdCost' in data[0][i]:
                data[0][i] = str(data[0][i]).replace('AdCost', 'Расход')

        for i in range(len(data)):
            for j in range(len(data[i])):
                wb.cell(row=i + 1, column=j + 1).value = data[i][j]

                if opt_symbol in data[0][j] and i != 0:
                    if type(data[i][j]) == type(0.0) or type(data[i][j]) == type(1):
                        continue
                    if "+" in data[i][j]:
                        wb.cell(row=i + 1, column=j + 1).fill = PatternFill(fill_type='solid', start_color='06d6a0',
                                                                            end_color='2ADEE9')
                    elif "-" in data[i][j]:
                        wb.cell(row=i + 1, column=j + 1).fill = PatternFill(fill_type='solid', start_color='ef476f',
                                                                            end_color='FC5977')
