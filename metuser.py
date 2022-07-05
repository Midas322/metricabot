#! /usr/bin/env python
# -*- coding: utf-8 -*-

from metmain import metmain
from openpyxl import Workbook
from CONSTANTS.DIR_SET import DIR_SET

class metuser(object):
    def __init__(self, met_id, targets, token, dir_client):
        self.met_id = met_id
        self.targets = targets
        self.token = token
        self.dir_client = dir_client
        self.errors = {}

    def make_file(self, _dates):
        met = metmain(self.met_id, self.token, self.targets,self.dir_client)
        wb = Workbook()

        dates = {
            'date1': {
                "{}".format(_dates[0])
            },
            'date2': {
                "{}".format(_dates[1])
            }
        }
        error_list = []
        for preset in DIR_SET.presets:
            try:
                if preset['dir_client'] and self.dir_client is not None and self.dir_client != "-":
                    data = met.make_data(preset['settings'], dates)
                elif not preset['dir_client']:
                    data = met.make_data(preset['settings'], dates)
                else:
                    continue
                data = met.cals_e_baisic(data)

                for dems in preset['dimensions']:
                    data = met.calc_efficiency_by_group(data, dems['dimensions'], dems['name'])
                met.make_page(met.make_list(data), preset['name'], wb)

                error_list.extend(met.errors)
            except:
                error_list.extend(met.errors)

        self.errors = list(set(error_list))
        print("Out Errors" + str(self.errors))

        sheets_names = wb.get_sheet_names()
        wb.remove(wb[sheets_names[0]])
        wb.save("documents/{}.xlsx".format(self.met_id))
        return "documents/{}.xlsx".format(self.met_id)

    def make_file_platforms(self, _dates):
        met = metmain(self.met_id, self.token, self.targets,self.dir_client)
        wb = Workbook()

        dates = {
            'date1': {
                "{}".format(_dates[0])
            },
            'date2': {
                "{}".format(_dates[1])
            }
        }
        error_list = []
        for preset in DIR_SET.presets_platforms:
            try:
                if preset['dir_client'] and self.dir_client is not None and self.dir_client != "-":
                    data = met.make_data(preset['settings'], dates)
                elif not preset['dir_client']:
                    data = met.make_data(preset['settings'], dates)
                else:
                    continue
                data = met.cals_e_baisic(data)

                for dems in preset['dimensions']:
                    data = met.calc_efficiency_by_group(data, dems['dimensions'], dems['name'])
                met.make_page(met.make_list(data), preset['name'], wb)

                error_list.extend(met.errors)
            except:
                error_list.extend(met.errors)

        self.errors = list(set(error_list))
        print("Out Errors" + str(self.errors))

        sheets_names = wb.get_sheet_names()
        wb.remove(wb[sheets_names[0]])
        wb.save("documents/{}.xlsx".format(self.met_id))
        return "documents/{}.xlsx".format(self.met_id)

    def make_file_by_time(self, _dates):
        met = metmain(self.met_id, self.token, self.targets, self.dir_client)
        wb = Workbook()

        dates = {
            'date1': {
                "{}".format(_dates[0])
            },
            'date2': {
                "{}".format(_dates[1])
            }
        }
        error_list = []
        for preset in DIR_SET.presets_by_time:
            print(preset['name'])
            try:
                if preset['dir_client'] and self.dir_client is not None and self.dir_client != "-":
                    data = met.make_data_bytime(preset['settings'], dates)
                elif not preset['dir_client']:
                    data = met.make_data_bytime(preset['settings'], dates)
                else:
                    continue
                data = met.cals_e_baisic(data)
                data = met.sum_days(data)
                data = met.calc_efficiency_by_group(data, [0], '')

                met.make_page(met.make_list(data), preset['name'], wb)

                error_list.extend(met.errors)
            except:
                error_list.extend(met.errors)

        self.errors = list(set(error_list))
        print("Out Errors" + str(self.errors))

        sheets_names = wb.get_sheet_names()
        wb.remove(wb[sheets_names[0]])
        wb.save("documents/{}.xlsx".format(self.met_id))
        return "documents/{}.xlsx".format(self.met_id)

    def quick_stat(self, _dates):
        met = metmain(self.met_id, self.token, self.targets, self.dir_client)

        dates = {
            'date1': [
                "{}".format(_dates[0])
            ],
            'date2': [
                "{}".format(_dates[1])
            ]
        }

        try:
            data = met.make_data(DIR_SET.quick_stat, dates)
            data = met.cals_e_baisic(data)[0]
        except:
            self.errors.append("Данные не получены. Возможно логин от Директа не Является главным или указан неверно. Подробнее смотрите тут: http://shendrik.site/bot")
            message = "Данные из Яндекс Директа не получены. Возможно логин от Директа не является главным или указан неверно. Подробнее смотрите тут: http://shendrik.site/bot"
            return message
        message=''
        message += 'Номер счетчика : {}\n'.format(self.met_id)
        message += 'Даты отчета : c {} по {}\n'.format(_dates[0], _dates[1])

        message += 'Расход Яндекс Директа : {:0.2f} руб. (Без НДС)\n'.format(data['metrics'][' ym:ad:<currency>AdCost'])
        message += 'Клики : {:0.0f}\n'.format(data['metrics']['ym:ad:visits'])
        message += 'CPC : {:0.2f} руб. (Без НДС)\n'.format(data['efficiency']['CPC'])

        if 'goals' in data['efficiency']:
            message += 'Заявки : {:0.0f}\n'.format(data['efficiency']['goals'])
            message += 'Конверсия : {:0.3f}%\n'.format(data['efficiency']['CR']*100)

        if 'CPA' in data['efficiency']:
            message += 'CPA : {:0.2f} руб. (Без НДС)\n'.format(data['efficiency']['CPA'])
        return message


if __name__ == '__main__':
    # Марат
    marat = metuser(met_id=45174921, targets=[54596614],
                    token='AgAAAAAgA9cWAAZHGdPwbQvdYUc1kWqo5hlAJS8', dir_client="Marat.direckt")

    marat.make_file(['2020-01-01', '2020-10-01'])
    # marat.make_file_by_time(['2020-01-01', '2020-10-01'])

    # marat.quick_stat(['2020-07-01','2020-09-01'])

